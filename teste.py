import xml.etree.ElementTree as ET
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font

# Caminho para o arquivo XML
file_path_xml = 'meuxml.xml'
file_path_xlsx = 'dados_nota_fiscal.xlsx'

# Carregar o XML
tree = ET.parse(file_path_xml)
root = tree.getroot()

# Namespaces
ns = {'nfe': 'http://www.portalfiscal.inf.br/nfe', 'ds': 'http://www.w3.org/2000/09/xmldsig#'}

# Função para buscar texto de um elemento XML com namespace
def find_text(element, tag, namespace):
    el = element.find(f"{namespace}:{tag}", ns)
    return el.text if el is not None else None

# Função para converter string para float, tratando valores que não podem ser convertidos
def safe_float(text):
    try:
        return float(text.replace(',', '.'))
    except (TypeError, ValueError):
        return 0.0

# Extraindo informações principais do XML
def extract_info_xml(root):
    products = defaultdict(lambda: {'xProd': '', 'qCom': 0.0, 'vUnCom': 0.0, 'vProd': 0.0})
    total_items = 0

    for det in root.findall('.//nfe:det', ns):
        total_items += 1
        prod = det.find('nfe:prod', ns)
        cProd = find_text(prod, 'cProd', 'nfe')
        xProd = find_text(prod, 'xProd', 'nfe')
        qCom = safe_float(find_text(prod, 'qCom', 'nfe'))
        vUnCom = safe_float(find_text(prod, 'vUnCom', 'nfe'))
        vProd = safe_float(find_text(prod, 'vProd', 'nfe'))

        key = (cProd, vUnCom)  # Chave composta por código e preço
        products[key]['xProd'] = xProd
        products[key]['qCom'] += qCom
        products[key]['vUnCom'] = vUnCom
        products[key]['vProd'] += vProd

    return products

# Extraindo informações do arquivo Excel
def extract_info_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    products = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        cProd = row[0]
        codigo_interno = row[-1]
        products[cProd] = codigo_interno

    return products

# Carregar informações do XML
xml_products = extract_info_xml(root)

# Carregar informações do arquivo Excel
xlsx_products = extract_info_xlsx(file_path_xlsx)

# Exportar os dados atualizados para um novo arquivo Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Dados da Nota Fiscal"

# Escrevendo cabeçalhos
headers = ["Codigo", "Nome do Produto", "Quantidade", "Valor Unitario", "Valor Total", "Codigo Interno"]
ws.append(headers)

# Definindo estilo para cabeçalhos
for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).font = Font(bold=True)

# Escrevendo dados dos produtos
for (cProd, vUnCom), info in xml_products.items():
    codigo_interno = xlsx_products.get(cProd, '')
    ws.append([cProd, info['xProd'], info['qCom'], info['vUnCom'], info['vProd'], codigo_interno])

# Salvando o arquivo Excel
wb.save('dados_atualizados.xlsx')
print("Dados exportados para dados_atualizados.xlsx")
